"""Excel (.xlsx) Parser. See ADR-0012.

`openpyxl` is loaded lazily inside `parse()` so the dep stays optional.
"""

import datetime as _datetime
import sys
from typing import Any, Dict, Iterator, Optional, Union

from filedge.parser import Parser


SheetSelector = Union[str, int, None]


class ExcelParser(Parser):
    mode = "rb"

    def __init__(self, sheet: SheetSelector = None):
        self._sheet = sheet

    def parse(self, fileobj) -> Iterator[Dict[str, Any]]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Excel support requires openpyxl — run: uv sync --extra excel"
            )

        wb = openpyxl.load_workbook(fileobj, read_only=True, data_only=True)
        ws = _select_sheet(wb, self._sheet)
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return
        headers = [str(h) if h is not None else "" for h in header_row]
        for row in rows:
            yield {headers[i]: _coerce(v) for i, v in enumerate(row) if i < len(headers)}


def _select_sheet(workbook, sheet: SheetSelector):
    names = workbook.sheetnames
    if sheet is None:
        if len(names) > 1:
            print(
                f"warning: workbook has {len(names)} sheets {names!r}; "
                f"reading first sheet {names[0]!r}. "
                f"Pass --sheet to choose a different sheet.",
                file=sys.stderr,
            )
        return workbook[names[0]]
    if isinstance(sheet, bool):
        raise TypeError(f"sheet must be str, int, or None — got bool {sheet!r}")
    if isinstance(sheet, int):
        if sheet < 0 or sheet >= len(names):
            raise ValueError(
                f"Sheet index {sheet} out of range; workbook has {len(names)} sheet(s): {names!r}"
            )
        return workbook[names[sheet]]
    if isinstance(sheet, str):
        if sheet not in names:
            raise ValueError(
                f"Missing sheet {sheet!r}; workbook has {names!r}"
            )
        return workbook[sheet]
    raise TypeError(f"sheet must be str, int, or None — got {type(sheet).__name__}")


def _coerce(value: Any) -> Optional[str]:
    if value is None:
        return None
    # bool MUST be checked before int — bool is a subclass of int in Python.
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, _datetime.datetime):
        return value.isoformat()
    if isinstance(value, _datetime.date):
        return value.isoformat()
    return str(value)
